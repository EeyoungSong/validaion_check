import pandas as pd
import os
import time
import zipfile
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# 감시할 폴더 경로
WATCH_FOLDER = r"C:\XBRL"

def is_valid_excel(file_path):
    """Excel 파일이 정상적인 ZIP 구조인지 확인"""
    return zipfile.is_zipfile(file_path)

def validate_and_highlight_excel(file_path, output_file):
    try:

        # 🔹 파일이 완전히 저장될 때까지 기다림
        time.sleep(5)

        # 🔹 엑셀 파일 읽기
        df = pd.read_excel(file_path, sheet_name=0, engine='openpyxl', dtype=str)
        df = df.astype(str).applymap(lambda x: x.strip() if isinstance(x, str) else x)

        # 🔹 필요한 열 데이터 추출
        sheet_col = df['시트'].tolist()
        mapping_col = df['맵핑표 번호'].tolist()
        xbrl_col = df['XBRL표 번호'].tolist()
        content_col = df['content'].tolist()
        depth_col = df['depth'].tolist()
        label_col = df['label(국문)'].tolist()

        error_cells = []  # 오류 발생한 셀 목록

        # ✅ **[1] 시트가 변경될 때 맵핑표 번호가 1로 리셋되는지 확인**
        for i in range(1, len(sheet_col)):  
            current_sheet, prev_sheet = str(sheet_col[i]), str(sheet_col[i-1])
            current_mapping, prev_mapping = mapping_col[i], mapping_col[i-1]

            if current_sheet != prev_sheet and current_mapping.strip() != "1":
                print("🚨 시트가 변경될 때 맵핑표 번호가 1로 리셋되는지 확인")
                error_cells.append((i + 2, df.columns.get_loc('맵핑표 번호') + 1))


        # ✅ **[2] 맵핑표 번호가 변경될 때 1씩 증가하는지 확인**
        for i in range(1, len(mapping_col)):  
            current_mapping, prev_mapping = mapping_col[i], mapping_col[i-1]
            current_sheet, prev_sheet = str(sheet_col[i]), str(sheet_col[i-1])
            if current_mapping != prev_mapping and current_sheet == prev_sheet:
                try:
                    if depth_col[i] != "1" and int(current_mapping) != int(prev_mapping) + 1:
                        print("🚨 맵핑 표번호 순서대로 증가하는지 확인")
                        error_cells.append((i + 2, df.columns.get_loc('맵핑표 번호') + 1))
                except ValueError:
                    print(f"⚠️ 숫자로 변환할 수 없는 맵핑 표번호: {current_mapping} 또는 {prev_mapping}")


        # ✅ **[3] XBRL 번호 변경 시 depth가 1로 초기화되는지 확인**
        prev_xbrl = xbrl_col[0].strip()
        for i in range(1, len(xbrl_col)):  
            current_mapping, prev_mapping = mapping_col[i], mapping_col[i-1]
            current_xbrl, current_depth = xbrl_col[i].strip(), depth_col[i].strip()
            current_xbrl_split = int(xbrl_col[i].strip().split(',')[0]) if ',' in xbrl_col[i] else int(xbrl_col[i])
            prev_xbrl_split = int(xbrl_col[i-1].strip().split(',')[1]) if ',' in xbrl_col[i-1] else int(xbrl_col[i-1])
            # if current_xbrl != prev_xbrl and current_depth != "1":
            #     print("🚨 XBRL 번호 변경 시 depth가 1로 초기화되는지 확인")
            #     error_cells.append((i + 2, df.columns.get_loc('depth') + 1))

            if current_mapping > prev_mapping and current_xbrl != prev_xbrl and current_xbrl_split != 0 and prev_xbrl_split != 0 and current_xbrl_split != prev_xbrl_split + 1:
                print("🚨 XBRL 번호 변경 시 이전 + 1인지 확인")
                error_cells.append((i + 2, df.columns.get_loc('XBRL표 번호') + 1))
            prev_xbrl = current_xbrl  



        # 📌 **오류가 발생한 셀을 노란색으로 하이라이트**
        wb = load_workbook(file_path)
        ws = wb.active
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for row, col in error_cells:
            ws.cell(row=row, column=col).fill = yellow_fill  # 해당 셀을 노란색으로 변경
            ws.cell(row=1, column=col).fill = yellow_fill  # 헤더도 색칠

        # 변경된 엑셀 저장
        wb.save(output_file)
        wb.close()

        # 오류 발생한 셀 출력
        if error_cells:
            print("🚨 오류 셀 목록:")
            for cell in error_cells:
                print(cell)
        else:
            print("✅ 오류 없음")

        return len(error_cells) == 0  # 오류가 없으면 True, 있으면 False
    
    except Exception as e:
        print(f"❌ 파일 검증 중 오류 발생: {e}")
        return False


class FileHandler(FileSystemEventHandler):
    def on_created(self, event):
        if not event.is_directory and event.src_path.lower().endswith(('.xlsx', '.xls')):
            file_path = event.src_path
            print(f"📂 새 파일 감지됨: {file_path}")

            # 현재 시간 기반으로 새로운 검증 파일명 생성
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            # 파일 검증 실행
            result = validate_and_highlight_excel(file_path, f"검증결과_{timestamp}.xlsx")

            if result:
                print(f"✅ 검증 성공! 파일 저장됨: {f"검증결과_{timestamp}.xlsx"}")
            else:
                print(f"❌ 검증 실패: {f"검증결과_{timestamp}.xlsx"}에 저장됨")

            # # 검증 완료 후 자동으로 파일 열기
            # try:
            #     print("📂 엑셀 파일을 열고 있습니다...")
            #     os.startfile(output_file)  
            # except Exception as e:
            #     print(f"❌ 파일 열기 실패: {e}")


if __name__ == "__main__":
    event_handler = FileHandler()
    observer = Observer()
    observer.schedule(event_handler, WATCH_FOLDER, recursive=False)
    observer.start()
    
    print(f"📡 폴더 모니터링 시작: {WATCH_FOLDER}")

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
        print("🛑 폴더 모니터링 중지됨")

    observer.join()
